
import sqlite3
from pathlib import Path
import openpyxl

DB = Path("grades.db")
XLSX = Path("副本培养计划课程安排.xlsx")

semester_blocks = [
    ("第一学期", 1, 4, 17),
    ("第三学期", 5, 4, 17),
    ("第五学期", 9, 4, 17),
    ("第七学期", 13, 4, 17),
    ("第二学期", 1, 21, 32),
    ("第四学期", 5, 21, 32),
    ("第六学期", 9, 21, 32),
    ("第八学期", 13, 21, 32),
]

conn = sqlite3.connect(DB)
cur = conn.cursor()

cur.execute("""
CREATE TABLE IF NOT EXISTS courses (
    id INTEGER PRIMARY KEY,
    semester TEXT NOT NULL,
    type TEXT NOT NULL,
    name TEXT NOT NULL,
    credit REAL NOT NULL,
    score REAL
)
""")

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["人工智能"]

course_id = 1
for semester, col, r1, r2 in semester_blocks:
    last_type = None
    for r in range(r1, r2 + 1):
        ctype = ws.cell(r, col).value
        name = ws.cell(r, col + 1).value
        credit = ws.cell(r, col + 2).value
        if ctype is not None:
            last_type = str(ctype).strip()
        if name is None:
            continue
        cur.execute("""
            INSERT OR REPLACE INTO courses (id, semester, type, name, credit, score)
            VALUES (?, ?, ?, ?, ?, COALESCE((SELECT score FROM courses WHERE id=?), NULL))
        """, (
            course_id,
            semester,
            last_type or "",
            str(name).strip(),
            float(credit or 0),
            course_id
        ))
        course_id += 1

conn.commit()
conn.close()
print("课程导入完成")
